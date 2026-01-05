import streamlit as st
import pandas as pd
import io
from datetime import datetime, timedelta
from utils.db import fetch_table
from utils.auth import require_pin
require_pin()
from utils.ui import apply_global_layout
apply_global_layout()




def export_excel_pretty(df: pd.DataFrame) -> bytes:
    """
    Exporta Excel bonitinho:
    - header negrito com fundo cinza
    - auto filtro
    - freeze header
    - largura automática
    """
    output = io.BytesIO()

    # ✅ Excel não suporta datetimes com timezone
    if "Data" in df.columns:
        df = df.copy()
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce", utc=True).dt.tz_localize(None)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Entregas")
        ws = writer.sheets["Entregas"]

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto filter
        ws.auto_filter.ref = ws.dimensions

        # Estilo header
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F4F4F")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

    return output.getvalue()


st.title("📊 Relatórios — Entregas e Exportação")

# ==========================
# CARREGAMENTO
# ==========================
deliveries = fetch_table("deliveries", order="delivered_at")
families = fetch_table("families")
leaders = fetch_table("cell_leaders")
baskets = fetch_table("basket_types")
cells = fetch_table("cells")
supers = fetch_table("supervisors")

if not deliveries:
    st.info("Nenhuma entrega registrada ainda.")
    st.stop()

fam_map = {f["id"]: f for f in families}
leader_map = {l["id"]: l for l in leaders}
basket_map = {b["id"]: b for b in baskets}
cell_map = {c["id"]: c for c in cells}
super_map = {s["id"]: s for s in supers}

# ==========================
# MONTAGEM DOS DADOS
# ==========================
rows = []
for d in deliveries:
    fam = fam_map.get(d["family_id"])
    leader = leader_map.get(d["leader_id"])
    basket = basket_map.get(d["basket_type_id"])

    # ======= CÉLULA =======
    cell = None
    if fam and fam.get("cell_id"):
        cell = cell_map.get(fam["cell_id"])

    cell_name = cell["cell_name"] if cell else "-"
    cell_network = cell["network_name"] if cell else "-"

    # ======= SUPERVISOR (fallback: célula -> líder) =======
    supervisor = None

    # 1) tenta supervisor pela célula
    if cell and cell.get("supervisor_id"):
        supervisor = super_map.get(cell["supervisor_id"])

    # 2) se não tiver supervisor na célula, tenta pelo líder
    if not supervisor and leader and leader.get("supervisor_id"):
        supervisor = super_map.get(leader["supervisor_id"])

    supervisor_name = supervisor["name"] if supervisor else "-"
    supervisor_phone = supervisor["phone"] if supervisor else "-"

    rows.append({
        "Data": d["delivered_at"],
        "Representante": fam["representative_name"] if fam else "-",
        "Telefone Representante": fam["representative_phone"] if fam else "-",
        "Célula": cell_name,
        "Rede da Célula": cell_network,
        "Supervisor": supervisor_name,
        "Telefone Supervisor": supervisor_phone,
        "Líder": leader["name"] if leader else "-",
        "Telefone Líder": leader["phone"] if leader else "-",
        "Rede do Líder": leader["network_name"] if leader else "-",
        "Cesta": basket["name"] if basket else "-",
        "Quantidade": d["quantity"]
    })

df = pd.DataFrame(rows)

# ✅ Converte Data já removendo timezone (Excel friendly)
df["Data"] = pd.to_datetime(df["Data"], errors="coerce", utc=True).dt.tz_localize(None)

# ==========================
# FILTROS
# ==========================
st.subheader("🔎 Filtros")

col1, col2, col3 = st.columns(3)
with col1:
    start = st.date_input("Data inicial", value=datetime.now().date() - timedelta(days=30))
with col2:
    end = st.date_input("Data final", value=datetime.now().date())
with col3:
    network_cell = st.selectbox("Rede da Célula", ["(todas)"] + sorted(df["Rede da Célula"].unique().tolist()))

leader_filter = st.selectbox("Líder", ["(todos)"] + sorted(df["Líder"].unique().tolist()))
basket_filter = st.selectbox("Tipo de cesta", ["(todas)"] + sorted(df["Cesta"].unique().tolist()))
cell_filter = st.selectbox("Célula", ["(todas)"] + sorted(df["Célula"].unique().tolist()))
supervisor_filter = st.selectbox("Supervisor", ["(todos)"] + sorted(df["Supervisor"].unique().tolist()))

f = df[(df["Data"].dt.date >= start) & (df["Data"].dt.date <= end)]

if network_cell != "(todas)":
    f = f[f["Rede da Célula"] == network_cell]
if leader_filter != "(todos)":
    f = f[f["Líder"] == leader_filter]
if basket_filter != "(todas)":
    f = f[f["Cesta"] == basket_filter]
if cell_filter != "(todas)":
    f = f[f["Célula"] == cell_filter]
if supervisor_filter != "(todos)":
    f = f[f["Supervisor"] == supervisor_filter]

# ==========================
# RESULTADOS
# ==========================
st.subheader("📋 Resultados")
st.dataframe(f.sort_values("Data", ascending=False), use_container_width=True)

# ==========================
# EXPORTAÇÃO
# ==========================
st.subheader("📥 Exportação")

# ✅ CSV padrão Excel BR
csv_data = f.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")

st.download_button(
    "⬇️ Baixar CSV (Excel - separado por ;)",
    csv_data,
    file_name="relatorio_entregas.csv",
    mime="text/csv"
)

# ✅ Excel bonito e profissional
xlsx_data = export_excel_pretty(f)

st.download_button(
    "⬇️ Baixar Excel (.xlsx) — Profissional",
    xlsx_data,
    file_name="relatorio_entregas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ==========================
# RESUMO
# ==========================
st.divider()
st.subheader("📌 Resumo")

colA, colB, colC = st.columns(3)
colA.metric("Entregas", len(f))
colB.metric("Total de cestas entregues", int(f["Quantidade"].sum()) if len(f) > 0 else 0)
colC.metric("Famílias atendidas", f["Telefone Representante"].nunique() if len(f) > 0 else 0)
